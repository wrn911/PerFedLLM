import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side


def convert_json_to_excel(json_file_path, output_excel_path):
    """
    将JSON数据转换为指定格式的Excel文件

    参数:
    json_file_path: JSON文件路径
    output_excel_path: 输出Excel文件路径
    """

    # 读取JSON数据
    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 转换为DataFrame
    df = pd.DataFrame(data)

    # --- 根据提供的JSON文件调整定义 ---

    # 1. 从数据中获取模型顺序，但尽量保持逻辑
    preferred_order = [
        "arima", "lasso", "svr", "lstm", "tft", 'autoformer',
        'dlinear', 'informer', 'simpletimellm', 'timellm'
    ]
    models_in_json = list(df['model'].unique())
    model_order = [m for m in preferred_order if m in models_in_json]
    model_order.extend([m for m in models_in_json if m not in model_order])

    # 2. 定义JSON中存在的城市顺序
    city_order = ['zte4gup', 'zte4gdown', 'zte5gup', 'zte5gdown']

    # 3. 数据集固定为 'sub'
    dataset_name = 'sub'

    # 4. 训练方法顺序（与您的代码一致）
    training_method_order = ['']
    # 训练方法的显示名称映射
    training_method_display = {
        '': 'Baseline',
        'fedavg': 'FedAvg',
        'fedprox': 'FedProx',
        'perfedavg': 'PerFedAvg'
    }

    # --- 创建Excel工作簿 ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # --- 设置表头 ---

    # 第一行：模型、训练方法、城市名称
    ws.cell(row=1, column=1, value="Model")
    ws.cell(row=1, column=2, value="Training Method")

    col_idx = 3
    for city in city_order:
        ws.cell(row=1, column=col_idx, value=city)
        # 合并每个城市的MSE和MAE列
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 1)
        col_idx += 2

    # 第二行：指标 (MSE/MAE)
    ws.cell(row=2, column=1, value="")  # A2
    ws.cell(row=2, column=2, value="")  # B2

    col_idx = 3
    for _ in city_order:
        ws.cell(row=2, column=col_idx, value="MSE")
        ws.cell(row=2, column=col_idx + 1, value="MAE")
        col_idx += 2

    # 合并A1:A2 和 B1:B2
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')

    # --- 填充数据 ---
    current_row = 3

    for model in model_order:
        model_start_row = current_row

        for tm_idx, training_method in enumerate(training_method_order):
            # 第一列：模型名称（只在每个模型的第一行显示）
            # （合并单元格将在数据填充后进行）
            if tm_idx == 0:
                ws.cell(row=current_row, column=1, value=model.capitalize())

            # 第二列：训练方法
            ws.cell(row=current_row, column=2, value=training_method_display[training_method])

            # 填充数据 (MSE/MAE)
            col_idx = 3  # 从第3列开始
            for city in city_order:
                for metric in ['mse', 'mae']:
                    # 查找对应的数据
                    filtered = df[
                        (df['model'] == model) &
                        (df['city'] == city) &
                        (df['dataset'] == dataset_name) &
                        (df['training_method'] == training_method)
                        ]

                    if not filtered.empty:
                        # 处理JSON中的重复条目（例如simpletimellm）
                        # 计算均值和标准差的均值
                        mean_val = filtered[f'{metric}_mean'].mean()
                        std_val = filtered[f'{metric}_std'].mean()

                        # 格式化数值为 "mean ± std"
                        if std_val == 0:
                            # 如果 std 为 0，只显示 mean
                            formatted_value = f"{mean_val:.4f}"
                        else:
                            formatted_value = f"{mean_val:.4f} ± {std_val:.4f}"

                        ws.cell(row=current_row, column=col_idx, value=formatted_value)
                    else:
                        ws.cell(row=current_row, column=col_idx, value="N/A")

                    col_idx += 1

            current_row += 1

    # --- 格式化 ---

    max_col = 2 + (len(city_order) * 2)

    # 合并模型单元格 (A列)
    model_start_row = 3
    num_tm_rows = len(training_method_order)
    if num_tm_rows > 1:
        for model in model_order:
            ws.merge_cells(
                start_row=model_start_row,
                start_column=1,
                end_row=model_start_row + num_tm_rows - 1,
                end_column=1
            )
            model_start_row += num_tm_rows

    # 设置列宽
    ws.column_dimensions['A'].width = 18  # Model列
    ws.column_dimensions['B'].width = 15  # Training Method列
    for col in range(3, max_col + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 18  # 数据列

    # 设置对齐方式和边框
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=1, max_row=current_row - 1, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border

    # 保存文件
    wb.save(output_excel_path)
    print(f"Excel文件已成功创建: {output_excel_path}")


# 使用示例
if __name__ == "__main__":
    # 确保 'statistical_results.json' 文件在同一目录下
    json_file_path = 'statistical_results.json'
    output_excel_path = 'statistical_results_formatted.xlsx'  # 更改了输出文件名

    convert_json_to_excel(json_file_path, output_excel_path)