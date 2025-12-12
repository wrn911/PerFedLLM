import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side


def convert_json_to_excel(json_file_path, output_excel_path):
    """
    将JSON数据转换为指定格式的Excel文件

    参数:
    json_file_path: JSON文件路径
    output_excel_path: 输出Excel文件路径
    """

    # 读取JSON数据
    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 转换为DataFrame
    df = pd.DataFrame(data)

    # 定义顺序
    model_order = ["arima", "lasso", "svr", "lstm", "tft", 'autoformer', 'dLinear', 'informer']
    dataset_order = ['call', 'net', 'sms']
    training_method_order = ['']

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # 设置表头
    # 第一行：城市名称
    ws.cell(row=1, column=1, value="")
    ws.cell(row=1, column=2, value="")
    ws.cell(row=1, column=3, value="Milano")
    ws.cell(row=1, column=9, value="Trento")

    # 第二行：指标类型
    ws.cell(row=2, column=1, value="")
    ws.cell(row=2, column=2, value="")
    ws.cell(row=2, column=3, value="MSE")
    ws.cell(row=2, column=6, value="MAE")
    ws.cell(row=2, column=9, value="MSE")
    ws.cell(row=2, column=12, value="MAE")

    # 第三行：数据集名称和训练方法标题
    ws.cell(row=3, column=1, value="Model")
    ws.cell(row=3, column=2, value="Training Method")
    datasets_labels = ['Call', 'Net', 'SMS']
    for i, dataset in enumerate(datasets_labels):
        ws.cell(row=3, column=3 + i, value=dataset)  # Milano MSE
        ws.cell(row=3, column=6 + i, value=dataset)  # Milano MAE
        ws.cell(row=3, column=9 + i, value=dataset)  # Trento MSE
        ws.cell(row=3, column=12 + i, value=dataset)  # Trento MAE

    # 合并单元格
    ws.merge_cells('C1:H1')  # Milano
    ws.merge_cells('I1:N1')  # Trento
    ws.merge_cells('C2:E2')  # Milano MSE
    ws.merge_cells('F2:H2')  # Milano MAE
    ws.merge_cells('I2:K2')  # Trento MSE
    ws.merge_cells('L2:N2')  # Trento MAE

    # 填充数据
    current_row = 4

    # 训练方法的显示名称映射
    training_method_display = {
        '': 'Baseline',
        'fedavg': 'FedAvg',
        'fedprox': 'FedProx',
        'perfedavg': 'PerFedAvg'
    }

    for model in model_order:
        model_start_row = current_row

        for tm_idx, training_method in enumerate(training_method_order):
            # 第一列：模型名称（只在第一行显示）
            if tm_idx == 0:
                ws.cell(row=current_row, column=1, value=model.capitalize())

            # 第二列：训练方法
            ws.cell(row=current_row, column=2, value=training_method_display[training_method])

            # 填充数据
            col = 3  # 从第3列开始
            for city in ['milano', 'trento']:
                for metric in ['mse', 'mae']:
                    for dataset in dataset_order:
                        # 查找对应的数据
                        filtered = df[
                            (df['model'] == model) &
                            (df['city'] == city) &
                            (df['dataset'] == dataset) &
                            (df['training_method'] == training_method)
                            ]

                        if not filtered.empty:
                            mean_val = filtered[f'{metric}_mean'].iloc[0]
                            std_val = filtered[f'{metric}_std'].iloc[0]

                            # 格式化数值
                            if mean_val < 1:
                                formatted_value = f"{mean_val:.6f} ± {std_val:.1f}"
                            else:
                                formatted_value = f"{mean_val:.6f} ± {std_val:.1f}"

                            ws.cell(row=current_row, column=col, value=mean_val)
                        else:
                            ws.cell(row=current_row, column=col, value="N/A")

                        col += 1

            current_row += 1

    # 设置列宽
    for col in range(1, 15):  # 增加到15列
        col_letter = get_column_letter(col)
        if col == 1:
            ws.column_dimensions[col_letter].width = 12  # Model列
        elif col == 2:
            ws.column_dimensions[col_letter].width = 15  # Training Method列
        else:
            ws.column_dimensions[col_letter].width = 18  # 数据列

    # 设置对齐方式
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 添加边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=1, max_row=current_row - 1, min_col=1, max_col=13):
        for cell in row:
            cell.border = thin_border

    # 保存文件
    wb.save(output_excel_path)
    print(f"Excel文件已成功创建: {output_excel_path}")


# 使用示例
if __name__ == "__main__":
    # 假设JSON数据保存在 'data.json' 文件中
    json_file_path = 'statistical_results.json'  # 您的JSON文件路径
    output_excel_path = 'results.xlsx'  # 输出Excel文件路径

    convert_json_to_excel(json_file_path, output_excel_path)